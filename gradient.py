from openpyxl import load_workbook
import shutil
import os
from openpyxl.styles import PatternFill

##################
# Getting Routes #
##################

def gradient_analysis(ruta):
    ruta_vehicular = os.path.join(ruta,"7.- Informacion de Campo","Vehicular")
    files = ["Tipico","Atipico"]
    listas = [[] for _ in range(2)]

    for index,file in enumerate(files):
        ruta_tipicidad = os.path.join(ruta_vehicular,file)
        try:
            list_excels = os.listdir(ruta_tipicidad)
        except FileNotFoundError:
            print(f"No hay datos en la carpeta de {file}")
            listas[index] = []
            list_excels = []

        for excel in list_excels:
            path_excel = os.path.join(ruta_tipicidad,excel)
            if not "~$" in excel:
                if file == "Tipico" and not '_GRAD' in excel:
                    listas[0].append(path_excel)
                elif file == "Atipico" and not '_GRAD' in excel:
                    listas[1].append(path_excel)

    for index,lista in enumerate(listas):
        if index == 0:
            print("#### Reportes Tipicos ####")
        else:
            print("#### Reportes Atipicos ####")
        for index2,ruta_excel in enumerate(lista):
            print(f"Analizando Excel ({index2+1}/{len(lista)})")
            wb = load_workbook(ruta_excel,read_only=True,data_only=True)
            hojas = ['N','S','E','O']
            giros = []
            list_columnas = []
            for hoja in hojas:
                ws = wb[hoja]
                giro = []
                #Encabezados
                for row in ws['K15:HB15']:
                    for elem in row:
                        valor = elem.value
                        if valor == 'N':
                            valor = None
                        giro.append(valor)
                giros.append(giro)

                #Datos por columnas
                columnas = [[] for _ in range(200)]
                for i,row in enumerate(ws['K16:HB111']):
                    for j in range(200):
                        valor = row[j].value
                        if type(valor) == str:
                            print("*****************************ERROR*******************************")
                            _, excel_nombre = os.path.split(ruta_excel)
                            if index == 0:
                                print("Carpeta: Tipico")
                            elif index==1:
                                print("Carpeta: Atipico")
                            print(f"Excel: {excel_nombre}")
                            print(f"Hoja: {hoja}")
                            print(f"Fila: {i+16} / Columna: {j+11}")
                            return print("*****************************ERROR*******************************")
                        if valor == None:
                            valor = 0
                        columnas[j].append(valor)
                list_columnas.append(columnas)

            wb.close()

            path_formato = "./tools/Formato_Vehicular.xlsx"
            final_path, name_excel = os.path.split(ruta_excel)
            new_directory = os.path.join(final_path+' (Protransito)','Gradiente')
            if not os.path.exists(new_directory):
                os.makedirs(new_directory)
            final_route = os.path.join(new_directory,name_excel[:-5]+"_GRAD.xlsx")
            shutil.copyfile(path_formato,final_route)

            wb = load_workbook(final_route)
            for index,hoja in enumerate(hojas):
                ws = wb[hoja]
                columns_to_hide = []
                for i,col in enumerate(list_columnas[index]):
                    if sum(col)==0:
                        columns_to_hide.append(11+i)
                    for j in range(96):
                        if j>=0:
                            celda = ws.cell(row=16+j,column=11+i)
                            if sum(col)==0:
                                celda.value = None
                            else:
                                celda.value = col[j]
                                if col[j-1] != 0:
                                    if (col[j]-col[j-1])/col[j-1]>1 or (col[j]-col[j-1])/col[j-1]<-1:
                                        relleno = PatternFill(start_color="FFFF0000", end_color="FFFF0000",fill_type="solid")
                                        celda.fill = relleno
                                    else:
                                        pass
                                if 24<=j<=37 or 48<=j<=59 or 70<=j<=81:
                                    if col[j]==0:
                                        celda.fill = PatternFill("solid", fgColor="FFFF00")

                for row in ws['K15:HB15']:
                    for i, elem in enumerate(row):
                        elem.value = giros[index][i]

                for col in columns_to_hide:
                    ws.column_dimensions[ws.cell(row=1,column=col).column_letter].hidden = True

            wb.save(final_route)
            wb.close()