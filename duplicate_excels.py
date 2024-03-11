from openpyxl import load_workbook, Workbook
import logging
import numpy as np
import os
import re
import time
from tqdm import tqdm

LOGGER =logging.getLogger(__name__)
LOGGER.setLevel(logging.DEBUG)
f = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

def read_excel_ped(excel_path) -> type(np.array): #type: ignore
    """Función que obtiene los conteos por giros de todos los peatonales para todos los sentidos."""
    turn_slice = [slice("G13","G22"),
                  slice("K13","K22"),
                  slice("G25","G34"),
                  slice("K25","K34")]
    
    matrix_slice = slice("L20","UY83")
    
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb['Inicio']

    try:
        quantities = [
            [row[0].value for row in ws[s]].index(None)
            for s in turn_slice
        ]
    except ValueError:
        quantities = []
        for s in turn_slice:
            column_data = [row[0].value for row in ws[s] if row[0].value is not None]
            if len(column_data) == 10:
                quantities.append(10)
            else:
                quantities.append(len(column_data))

    ws = wb['Data Peatonal']

    #Reading data from sheets
    list_A = [[cell.value for cell in row] for row in ws[matrix_slice]]    

    try:
        A = np.array(list_A, dtype="float")
    except TypeError:
        for i, row in enumerate(list_A):
            for j, value in enumerate(row):
                if isinstance(value, str):
                    LOGGER.error(f"Row: {i+19}, Column: {j+11}, Value: {value}")

    A[np.isnan(A)] = 0

    morning = []
    evening = []
    night   = []

    INCREMENTOS = [0,140,280,420]

    for incremento, giros in zip(INCREMENTOS, quantities):
        morning.append(
            np.array(
                [
                    A[2:14, (10*ped_type+incremento):(10*ped_type+giros+incremento)]
                    for ped_type in range(4) #Kid, Adult, Old, WC
                ]
            )
        )
        evening.append(
            np.array(
                [
                    A[24:36, (10*ped_type+incremento):(10*ped_type+giros+incremento)]
                    for ped_type in range(4) #Kid, Adult, Old, WC
                ]
            )
        )
        night.append(
            np.array(
                [
                    A[46:58, (10*ped_type+incremento):(10*ped_type+giros+incremento)]
                    for ped_type in range(4) #Kid, Adult, Old, WC
                ]
            )
        )

    wb.close()

    MORNING = []
    EVENING = []
    NIGHT = []

    for morning_i, evening_i, night_i in zip(morning, evening, night):
        morning_concatenated = np.concatenate(morning_i, axis = 1)
        evening_concatenated = np.concatenate(evening_i, axis = 1)
        night_concatenated = np.concatenate(night_i, axis = 1)
        MORNING.append(morning_concatenated)
        EVENING.append(evening_concatenated)
        NIGHT.append(night_concatenated)

    MORNING = np.concatenate(MORNING, axis = 1)
    EVENING = np.concatenate(EVENING, axis = 1)
    NIGHT = np.concatenate(NIGHT, axis = 1)
    
    EXCEL = []
    EXCEL.append([MORNING, EVENING, NIGHT])
    return EXCEL

def read_excel_veh(excel_path, num_veh_classes
               ) -> type(np.array) and type(np.array) and list: #type: ignore
    """Función que obtiene los conteos por giros de todos los vehículos para todos los sentidos."""
    
    #Slices definition
    turn_slice = [slice("G12","G21"),
                  slice("M12","M21"),
                  slice("G24","G33"),
                  slice("M24","M33")]

    #Reading number of turns
    hojas = ['N','S','E','O']
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    LOGGER.info(f"Lectura de excel:\n{excel_path}")
    ws = wb['Inicio']

    try:
        quantities = [
            [row[0].value for row in ws[s]].index(None)
            for s in turn_slice
        ]
    except ValueError:
        quantities = []
        for s in turn_slice:
            column_data = [row[0].value for row in ws[s] if row[0].value is not None]
            if len(column_data) == 10:
                quantities.append(10)
            else:
                quantities.append(len(column_data))
        LOGGER.error(f"Hay algun dato que no es número en la base de datos")

    #Reading data from sheets
    EXCEL = []
    for (index, hoja), num in zip(enumerate(hojas),quantities):
        num_giro_i = quantities[index]
        ws = wb[hoja]
        list_A = [[cell.value for cell in row] for row in ws["K16":"HB111"]]
            
        try:
            A = np.array(list_A, dtype="float")
        except ValueError:
            #Revisión si hay textos.
            for i, row in enumerate(list_A):
                for j, value in enumerate(row):
                    if isinstance(value, str):
                        LOGGER.critical(f"Row: {i+16}, Column: {j+11}. Value = {value}, Hoja = {hoja}, Excel = {os.path.split(excel_path)[1]}")
                        return print("Se han encontrado valores que no son float o int, revisar LOGS/sheets.log")

        A[np.isnan(A)] = 0

        flow = np.concatenate(
            [
                A[:, (10*veh_type) : (10*veh_type + num_giro_i)]
                for veh_type in range(num_veh_classes)
            ],
            axis=1
        )

        MORNING     = flow[26:38,:]
        EVENING     = flow[48:60,:]
        NIGHT       = flow[70:82,:]

        EXCEL.append([MORNING, EVENING, NIGHT]) #POR HOJA

    wb.close()

    return EXCEL

def find_duplicate_excels(EXCEL_T, EXCEL_A, ws, accum_count, codigo_t, codigo_a) -> int:
    count = accum_count
    hojas = ["N","S","E","O"]
    STOP = False
    for i, sheets in enumerate(EXCEL_T):
        if STOP: break
        MORNING = sheets[0]
        EVENING = sheets[1]
        NIGHT = sheets[2]

        for j in range(MORNING.shape[1]):
            if STOP: break
            GIRO_M = MORNING[:,j] #MATRIZ NP.ARRAY COLUMNA POR TURNO.
            GIRO_E = EVENING[:,j]
            GIRO_N = NIGHT[:,j]
            for m, sht in enumerate(EXCEL_A):
                if STOP: break
                morning = sht[0]
                evening = sht[1]
                night = sht[2]
                for k in range(morning.shape[1]):
                    if STOP: break
                    giro_m = morning[:,k]
                    giro_e = evening[:,k]
                    giro_n = night[:,k]

                    repetition = 0
                    if np.array_equal(giro_m, GIRO_M):
                        repetition = GIRO_M
                        turno = "Mañana"
                    elif np.array_equal(giro_e, GIRO_E):
                        repetition = GIRO_E
                        turno = "Tarde"
                    elif np.array_equal(giro_n, GIRO_N):
                        repetition = GIRO_N
                        turno = "Noche"

                    if type(repetition) == int or sum(repetition) < 12: continue
                    else:
                        result = ""
                        for elem in repetition:
                            result += str(elem)+', '
                        ws.cell(row=count+2, column=1, value=result[:-2])
                        ws.cell(row=count+2, column=2, value=hojas[i])
                        ws.cell(row=count+2, column=3, value=hojas[m])
                        ws.cell(row=count+2, column=4, value=turno)
                        ws.cell(row=count+2, column=5, value=codigo_t)
                        ws.cell(row=count+2, column=6, value=codigo_a)
                        count += 1
                        STOP = True
                        break
    return count

def excels_duplicated_excels_veh(directory):
    vehicle_path = os.path.join(directory,"7.- Informacion de Campo", "Vehicular")
    tipico_files = os.listdir(os.path.join(vehicle_path,"Tipico"))
    atipico_files = os.listdir(os.path.join(vehicle_path,"Atipico"))

    logger_path = os.path.join(vehicle_path, "LOGS")
    if not os.path.exists(logger_path):
        os.mkdir(logger_path)
    fh = logging.FileHandler(os.path.join(vehicle_path, "LOGS", "excels_car.log"))
    fh.setFormatter(f)
    LOGGER.addHandler(fh)

    tipico_files = [file for file in tipico_files if file.endswith(".xlsm") and not file.startswith("~")]
    atipico_files = [file for file in atipico_files if file.endswith(".xlsm") and not file.startswith("~")]

    summary_comparison = os.path.join(vehicle_path, "Summary_Excels_Cars.xlsx")
    wb = Workbook()
    wb.save(summary_comparison)
    wb.close()

    wb = load_workbook(summary_comparison)
    ws = wb['Sheet']

    ws.cell(row=1, column=1, value="Patrón")
    ws.cell(row=1, column=2, value="Hoja 1")
    ws.cell(row=1, column=3, value="Hoja 2")
    ws.cell(row=1, column=4, value="Turno")
    ws.cell(row=1, column=5, value="Excel Tipico")
    ws.cell(row=1, column=6, value="Excel Atipico")

    pattern1 = r"([A-Z]+[0-9]+)"
    pattern2 = r"([A-Z]+-[0-9]+)"

    accum_count = 0
    print("###### STARTING EXCEL COMPARISONS ######")
    for i, tipico in tqdm(enumerate(tipico_files), desc="Procesando Tipico vs Atipico"):
        coincidense_t = re.search(pattern1, tipico)
        coincidense_t2 = re.search(pattern2, tipico)
        if coincidense_t:
            codigo_t = coincidense_t.group(1)
            codigo_t = codigo_t[:2] +'-' + codigo_t[2:]
        elif coincidense_t2:
            codigo_t = coincidense_t2.group(1)
        else:
            print(f"Para este excel no existe código: {tipico}")
            continue
        for atipico in atipico_files:
            coincidense_a = re.search(pattern1, atipico)
            coincidense_a2 = re.search(pattern2, atipico)
            if coincidense_a:
                codigo_a = coincidense_a.group(1)
                codigo_a = codigo_a[:2]+'-'+codigo_a[2:]
            elif coincidense_a2:
                codigo_a = coincidense_a2.group(1)
            else:
                print(f"Para el siguiente excel no existe código: {atipico}")
                continue
            route_tipico = os.path.join(vehicle_path,"Tipico", tipico)
            route_atipico = os.path.join(vehicle_path,"Atipico", atipico)

            EXCEL_TIPICO = read_excel_veh(route_tipico, 11)
            EXCEL_ATIPICO = read_excel_veh(route_atipico, 11)

            current_count = find_duplicate_excels(EXCEL_TIPICO, EXCEL_ATIPICO, ws, accum_count, codigo_t, codigo_a)
            accum_count = current_count

    wb.save(summary_comparison)
    wb.close()
    wb = load_workbook(summary_comparison)
    ws = wb['Sheet']

    for i, tipico1 in tqdm(enumerate(tipico_files), desc="Procesando Tipico vs Tipico"):
        coincidense_t1 = re.search(pattern1, tipico1)
        coincidense_t1_2 = re.search(pattern2, tipico1)
        if coincidense_t1:
            codigo_t1 = coincidense_t1.group(1)
            codigo_t1 = codigo_t1[:2]+'-'+codigo_t1[2:]
        elif coincidense_t1_2:
            codigo_t1 = coincidense_t1_2.group(1)
        else:
            print(f"Para este no existe código {tipico1}")
            continue
        for tipico2 in tipico_files:
            coincidense_t2 = re.search(pattern1, tipico2)
            coincidense_t2_2 = re.search(pattern2, tipico2)
            if coincidense_t2:
                codigo_t2 = coincidense_t2.group(1)
                codigo_t2 = codigo_t2[:2]+'-'+codigo_t2[2:]
            elif coincidense_t2_2:
                codigo_t2 = coincidense_t2_2.group(1)
            else:
                print(f"Para el siguiente excel no existe código: {tipico2}")
                continue

            route_tipico1 = os.path.join(vehicle_path, "Tipico", tipico1)
            route_tipico2 = os.path.join(vehicle_path, "Tipico", tipico2)
            
            EXCEL_TIPICO1 = read_excel_veh(route_tipico1, 11) #<---- Se puede cambiar el número de tipos de vehiculares a comparar
            EXCEL_TIPICO2 = read_excel_veh(route_tipico2, 11) #<---- Se puede cambiar el número de tipos de vehiculares a comparar

            current_count = find_duplicate_excels(EXCEL_TIPICO1, EXCEL_TIPICO2, ws, accum_count, codigo_t1, codigo_t2)

    wb.save(summary_comparison)
    wb.close()
    wb = load_workbook(summary_comparison)
    ws = wb['Sheet']

    for i, atipico1 in enumerate(atipico_files):
        print(f"Analizando excel ({i+1}/{len(atipico_files)})")
        coincidense_a1 = re.search(pattern1, atipico1)
        coincidense_a1_2 = re.search(pattern2, atipico1)
        if coincidense_a1:
            codigo_a1 = coincidense_a1.group(1)
            codigo_a1 = codigo_a1[:2]+'-'+codigo_a1[2:]
        elif coincidense_a1_2:
            codigo_a1 = coincidense_a1_2.group(1)
        else:
            print(f"Para el siguiente excel no existe código: {atipico1}")
            continue
        for atipico2 in atipico_files:
            coincidense_a2 = re.search(pattern1, atipico2)
            coincidense_a2_2 = re.search(pattern2, atipico2)
            if coincidense_a2:
                codigo_a2 = coincidense_a2.group(1)
                codigo_a2 = codigo_a2[:2]+'-'+codigo_a2[2:]
            elif coincidense_a2_2:
                codigo_a2 = coincidense_a2_2.group(1)
            else: 
                print(f"Para el siguiente excel no existe código: {atipico2}")
                continue
            route_atipico1 = os.path.join(vehicle_path, "Atipico", atipico1)
            route_atipico2 = os.path.join(vehicle_path, "Atipíco", atipico2)

            EXCEL_ATIPICO1 = read_excel_veh(route_atipico1, 11) #<---- Se puede cambiar el número de tipos de vehiculares a comparar
            EXCEL_ATIPICO2 = read_excel_veh(route_atipico2, 11) #<---- Se puede cambiar el número de tipos de vehiculares a comparar

            current_count = find_duplicate_excels(EXCEL_ATIPICO1, EXCEL_ATIPICO2, ws, accum_count, codigo_a1, codigo_a2)

    wb.save(summary_comparison)
    wb.close()

def excels_duplicated_excels_ped(directory):
    pedestrian_path = os.path.join(directory,"7.- Informacion de Campo", "Peatonal")
    tipico_files = os.listdir(os.path.join(pedestrian_path,"Tipico"))
    atipico_files = os.listdir(os.path.join(pedestrian_path,"Atipico"))

    logger_path = os.path.join(pedestrian_path, "LOGS")
    if not os.path.exists(logger_path):
        os.mkdir(logger_path)
    fh = logging.FileHandler(os.path.join(pedestrian_path, "LOGS", "excels_ped.log"))
    fh.setFormatter(f)
    LOGGER.addHandler(fh)

    tipico_files = [file for file in tipico_files if file.endswith(".xlsm") and not file.startswith("~")]
    atipico_files = [file for file in atipico_files if file.endswith(".xlsm") and not file.startswith("~")]

    summary_comparison = os.path.join(pedestrian_path, "Summary_Excels_Peds.xlsx")
    wb = Workbook()
    wb.save(summary_comparison)
    wb.close()

    wb = load_workbook(summary_comparison)
    ws = wb['Sheet']

    ws.cell(row=1, column=1, value="Patrón")
    ws.cell(row=1, column=2, value="Hoja 1")
    ws.cell(row=1, column=3, value="Hoja 2")
    ws.cell(row=1, column=4, value="Turno")
    ws.cell(row=1, column=5, value="Excel Tipico")
    ws.cell(row=1, column=6, value="Excel Atipico")

    pattern = r"([A-Z]+[0-9]+)"
    pattern2 = r"([A-Z]+-[0-9]+)"
    accum_count = 0

    start_time = time.time()
    for tipico in tqdm(tipico_files, desc = "Procesando Típico vs Atípico"):
        coincidense_t = re.search(pattern, tipico)
        coincidense_t2 = re.search(pattern2, tipico)
        if coincidense_t:
            codigo_t = coincidense_t.group(1)
        elif coincidense_t2:
            codigo_t = coincidense_t2.group(1)
        else:
            LOGGER.error(f"Para el siguiente excel no existe código: {tipico}")
            continue

        for atipico in atipico_files:
            coincidense_a = re.search(pattern, atipico)
            coincidense_a2 = re.search(pattern2, atipico)
            if coincidense_a:
                codigo_a = coincidense_a.group(1)
            elif coincidense_a2:
                codigo_a = coincidense_a2.group(1)
            else:
                LOGGER.error(f"Para el siguiente excel no existe código: {atipico}")
                continue

            route_tipico = os.path.join(pedestrian_path,"Tipico", tipico)
            route_atipico = os.path.join(pedestrian_path,"Atipico", atipico)

            try:
                EXCEL_TIPICO = read_excel_ped(route_tipico)
            except:
                LOGGER.error(f"Error en este excel:\n{tipico}")
                break
            try:
                EXCEL_ATIPICO = read_excel_ped(route_atipico)
            except:
                LOGGER.error(f"Error en este excel:\n{atipico}")
                continue

            current_count = find_duplicate_excels(EXCEL_TIPICO, EXCEL_ATIPICO, ws, accum_count, codigo_t, codigo_a)
            accum_count = current_count
    end_time = time.time()
    excel_time = end_time-start_time
    LOGGER.info(f"Time: {excel_time:.2f}")

    wb.save(summary_comparison)
    wb.close()

    wb = load_workbook(summary_comparison)
    ws = wb['Sheet']

    start_time = time.time()
    for tipico in tqdm(tipico_files, desc="Procesando Típico vs Típico"):
        coincidense_t_1 = re.search(pattern, tipico)
        coincidense_t2_1 = re.search(pattern2, tipico)
        if coincidense_t_1:
            codigo_t_1 = coincidense_t_1.group(1)
        elif coincidense_t2_1:
            codigo_t_1 = coincidense_t2_1.group(1)
        else:
            LOGGER.error(f"Para el siguiente excel no existe código: {tipico}")
            continue

        for tipico2 in tipico_files:
            coincidense_t_2 = re.search(pattern, tipico2)
            coincidense_t2_2 = re.search(pattern2, tipico2)
            if coincidense_t_2:
                codigo_t_2 = coincidense_t_2.group(1)
            elif coincidense_t2_2:
                codigo_t_2 = coincidense_t2_2.group(1)
            else:
                LOGGER.error(f"Para el siguiente excel no existe código: {tipico2}")
                continue

            route_tipico1 = os.path.join(pedestrian_path, "Tipico", tipico)
            route_tipico2 = os.path.join(pedestrian_path, "Tipico", tipico2)

            if codigo_t_1 == codigo_t_2: continue
            
            try:
                EXCEL_TIPICO_1 = read_excel_ped(route_tipico1)
            except:
                LOGGER.error(f"Error en este excel:\n{tipico}")
                break

            try:
                EXCEL_TIPICO_2 = read_excel_ped(route_tipico2)
            except:
                LOGGER.error(f"Error en este excel:\n{tipico2}")
                continue

            current_count = find_duplicate_excels(EXCEL_TIPICO_1, EXCEL_TIPICO_2, ws, accum_count, codigo_t_1, codigo_t_2)
            accum_count = current_count
            
    end_time = time.time()
    excel_time = end_time-start_time
    LOGGER.info(f"Time: {excel_time:.2f}")

    wb.save(summary_comparison)
    wb.close()

    wb = load_workbook(summary_comparison)
    ws = wb['Sheet']

    start_time = time.time()
    for atipico in tqdm(atipico_files, desc="Procesando Atípico vs Atípico"):
        coincidense_a_1 = re.search(pattern, atipico)
        coincidense_a2_1 = re.search(pattern2, atipico)
        if coincidense_a_1:
            codigo_a_1 = coincidense_a_1.group(1)
        elif coincidense_a2_1:
            codigo_a_1 = coincidense_a2_1.group(1)
        else:
            LOGGER.error(f"Para el siguiente excel no existe código: {atipico}")

        for atipico2 in atipico_files:
            coincidense_a_2 = re.search(pattern, atipico2)
            coincidense_a2_2 = re.search(pattern2, atipico2)
            if coincidense_a_2:
                codigo_a_2 = coincidense_a_2.group(1)
            if coincidense_a2_2:
                codigo_a_2 = coincidense_a2_2.group(1)
            else:
                LOGGER.error(f"Para el siguiente excel no existe código: {atipico2}")

            route_atipico1 = os.path.join(pedestrian_path, "Atipico", atipico)
            route_atipico2 = os.path.join(pedestrian_path, "Atipico", atipico2)

            if codigo_a_1 == codigo_a_2: continue

            try:
                EXCEL_ATIPICO_1 = read_excel_ped(route_atipico1)
            except:
                LOGGER.error(f"Error en este excel:\n{atipico}")
                break

            try:
                EXCEL_ATIPICO_2 = read_excel_ped(route_atipico2)
            except:
                LOGGER.error(f"Error en este excel:\n{atipico2}")
                continue

            current_count = find_duplicate_excels(EXCEL_ATIPICO_1, EXCEL_ATIPICO_2, ws, accum_count, codigo_a_1, codigo_a_2)
            accum_count = current_count
    end_time = time.time()
    excel_time = end_time-start_time
    LOGGER.info(f"Time: {excel_time:.2f}")

    wb.save(summary_comparison)
    wb.close()