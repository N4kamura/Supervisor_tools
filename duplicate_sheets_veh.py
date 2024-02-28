from openpyxl import load_workbook, Workbook
import logging
import numpy as np
import os

#Define Logger

LOGGER = logging.getLogger(__name__)
LOGGER.setLevel(logging.DEBUG)
f = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    
def read_excel_veh(excel_path, num_veh_classes
               ) -> type(np.array) and type(np.array) and list: #type: ignore
    """Función que obtiene los conteos por giros de todos los vehículos para todos los sentidos."""
    
    #Slices definition
    turn_slice = [slice("G12","G21"),
                  slice("M12","M21"),
                  slice("G24","G33"),
                  slice("M24","M33")]

    #Reading number of turns
    hojas = ['N','S','E','O']
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    LOGGER.info(f"Lectura de excel:\n{excel_path}")
    ws = wb['Inicio']

    try:
        quantities = [
            [row[0].value for row in ws[s]].index(None)
            for s in turn_slice
        ]
    except ValueError:
        quantities = []
        for s in turn_slice:
            column_data = [row[0].value for row in ws[s] if row[0].value is not None]
            if len(column_data) == 10:
                quantities.append(10)
            else:
                quantities.append(len(column_data))
        LOGGER.error(f"Hay algun dato que no es número en la base de datos")

    #Reading data from sheets
    EXCEL = []
    for (index, hoja), num in zip(enumerate(hojas),quantities):
        num_giro_i = quantities[index]
        ws = wb[hoja]
        list_A = [[cell.value for cell in row] for row in ws["K16":"HB111"]]
            
        try:
            A = np.array(list_A, dtype="float")
        except ValueError:
            #Revisión si hay textos.
            for i, row in enumerate(list_A):
                for j, value in enumerate(row):
                    if isinstance(value, str):
                        LOGGER.critical(f"Row: {i+16}, Column: {j+11}. Value = {value}, Hoja = {hoja}, Excel = {os.path.split(excel_path)[1]}")
                        return print("Se han encontrado valores que no son float o int, revisar LOGS/sheets.log")

        A[np.isnan(A)] = 0

        flow = np.concatenate(
            [
                A[:, (10*veh_type) : (10*veh_type + num_giro_i)]
                for veh_type in range(num_veh_classes)
            ],
            axis=1
        )

        MORNING     = flow[26:38,:]
        EVENING     = flow[48:60,:]
        NIGHT       = flow[70:82,:]

        EXCEL.append([MORNING, EVENING, NIGHT]) #POR HOJA

    wb.close()

    return EXCEL

def find_duplicate_by_sheets(EXCEL, ws, accum_count, excel_name) -> int:
    
    count = accum_count
    hojas = ["N", "S", "E", "O"]
    STOP = False
    for i, sheets in enumerate(EXCEL):
        if i == len(EXCEL)-1:
            break
        if STOP: break
        MORNING = sheets[0]
        EVENING = sheets[1]
        NIGHT   = sheets[2]
        for j in range(MORNING.shape[1]):
            if STOP: break
            GIRO_M = MORNING[:,j] #MATRIZ NP.ARRAY COLUMNA POR TURNO.
            GIRO_E = EVENING[:,j]
            GIRO_N = NIGHT[:,j]
            for m, sht in enumerate(EXCEL[i+1:]):
                if STOP: break
                morning = sht[0]
                evening = sht[1]
                night   = sht[2]
                for k in range(morning.shape[1]):
                    if STOP: break
                    giro_m = morning[:,k]
                    giro_e = evening[:,k]
                    giro_n = night[:,k]

                    repetition = 0
                    if np.array_equal(giro_m, GIRO_M):
                        repetition = GIRO_M
                        turno = "Mañana"
                    elif np.array_equal(giro_e, GIRO_E):
                        repetition = GIRO_E
                        turno = "Tarde"
                    elif np.array_equal(giro_n, GIRO_N):
                        repetition = GIRO_N
                        turno = "Noche"
                    
                    if type(repetition) == int or sum(repetition) < 12: continue
                    else:
                        result = ""
                        for elem in repetition:
                            result += str(elem)+', '
                        ws.cell(row=count+2, column=1, value=result[:-2])
                        ws.cell(row=count+2, column=2, value=hojas[i])
                        ws.cell(row=count+2, column=3, value=hojas[i+1+m])
                        ws.cell(row=count+2, column=4, value=turno)
                        ws.cell(row=count+2, column=5, value=excel_name)
                        count += 1
                        STOP = True
                        break
    return count

def sheets_duplicated(directory) -> None:
    vehicle_path = os.path.join(directory,"7.- Informacion de Campo", "Vehicular")
    tipico_files = os.listdir(os.path.join(vehicle_path,"Tipico"))
    atipico_files = os.listdir(os.path.join(vehicle_path,"Atipico"))

    logger_path = os.path.join(vehicle_path, "LOGS")
    if not os.path.exists(logger_path):
        os.mkdir(logger_path)
    fh = logging.FileHandler(os.path.join(vehicle_path, "LOGS", "sheets_cars.log"))
    fh.setFormatter(f)
    LOGGER.addHandler(fh)

    tipico_files = [file for file in tipico_files if file.endswith(".xlsm") and not file.startswith("~")]
    atipico_files = [file for file in atipico_files if file.endswith(".xlsm") and not file.startswith("~")]
    summary_excel = os.path.join(vehicle_path, "Summary_sheets_cars.xlsx")
    wb = Workbook()
    wb.save(summary_excel)
    wb.close()

    wb = load_workbook(summary_excel)
    ws = wb["Sheet"]

    ws.cell(row=1, column=1, value="Patrón")
    ws.cell(row=1, column=2, value="Hoja 1")
    ws.cell(row=1, column=3, value="Hoja 2")
    ws.cell(row=1, column=4, value="Turno")
    ws.cell(row=1, column=5, value="Excel")

    accum_count = 0
    print("####### STARTING SHEETS COMPARISONS ######")
    print("################ TIPICO ##################")
    for i, excel in enumerate(tipico_files):
        print(f"Analizando Excel ({i+1}/{len(tipico_files)})")
        route_excel = os.path.join(vehicle_path, "Tipico", excel)
        try:
            EXCEL = read_excel_veh(route_excel, 11)
            current_count = find_duplicate_by_sheets(EXCEL, ws, accum_count, excel)
            accum_count = current_count
        except Exception as e:
            print(f"Error en este excel:\n{excel}")
            LOGGER.error(f"Error en este excel:\n{excel}")
        
    print("############### ATIPICO ###################")
    for i, excel in enumerate(atipico_files):
        print(f"Analizando Excel ({i+1}/{len(atipico_files)})")
        route_excel = os.path.join(vehicle_path, "Atipico", excel)
        try:
            EXCEL = read_excel_veh(route_excel, 11)
            current_count = find_duplicate_by_sheets(EXCEL, ws, accum_count, excel)
            accum_count = current_count
        except Exception as e:
            print(f"Error en este excel:\n{excel}")
            LOGGER.error(f"Error en este excel:\n{excel}")

    LOGGER.info(f"Se han encontrado {accum_count} coincidencias")
        
    wb.save(summary_excel)
    wb.close()