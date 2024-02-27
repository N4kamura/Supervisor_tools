from openpyxl import load_workbook, Workbook
import os
import re

def _reading_dates(path_excel, letter) -> str:
    pattern1 = r"([A-Z]+-[0-9]+)"
    pattern2 = r"([A-Z]+[0-9]+)"
    name_excel = os.path.split(path_excel)[1]
    coincidence = re.search(pattern1, name_excel)
    coincidence2 = re.search(pattern2, name_excel)
    if coincidence:
        code = coincidence.group(1)
    elif coincidence2:
        code = coincidence2.group(1)
        code = code[:2]+'-'+code[2:4]
    else: print(f"No hay coincidencias en el excel:\n{name_excel}")

    wb = load_workbook(path_excel, read_only=True, data_only=True)
    ws = wb['Inicio']
    date = ws[letter].value
    wb.close()
    try:
        date = date.strftime('%Y-%m-%d')
    except AttributeError as e:
        date = "ERROR"

    return date, code

def find_files(main_path) -> list:
    data = []
    field_data = os.path.join(main_path, '7. Informacion de Campo')
    type_data = os.listdir(field_data)
    type_data = [tipo for tipo in type_data if "Tiempo" not in tipo]
    tipicidades = ["Tipico", "Atipico"]
    for tipo in type_data:
        print(f"#### {tipo} ####")
        type_route = os.path.join(field_data, tipo)
        if tipo == "Vehicular": letter = "G6"
        elif tipo == "Peatonal": letter = "G5"
        else: continue

        for tipicidad in tipicidades:
            print(f"{tipicidad}:")
            tipicidad_route = os.path.join(type_route, tipicidad)
            files = os.listdir(tipicidad_route)
            files = [file for file in files if "~$" not in file and ".xlsm" in file]
            for j, file in enumerate(files):
                print(f"Excel {j+1}/{len(files)}")
                path_excel = os.path.join(tipicidad_route, file)
                date, code = _reading_dates(path_excel, letter)
                data.append((code, date, tipicidad, tipo))

    return data

def create_summary(data, main_path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Código")
    ws.cell(row=1, column=2, value="Fecha")
    ws.cell(row=1, column=3, value="Tipicidad")
    ws.cell(row=1, column=4, value="Tipo")
    for i, dt in enumerate(data):
        ws.cell(row=i+2, column=1, value=dt[0])
        ws.cell(row=i+2, column=2, value=dt[1])
        ws.cell(row=i+2, column=3, value=dt[2])
        ws.cell(row=i+2, column=4, value=dt[3])

    wb.save(os.path.join(main_path, '7. Informacion de Campo', 'Date_Summary.xlsx'))
    wb.close()
    print("Finalizado")