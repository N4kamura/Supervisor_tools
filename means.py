from openpyxl import load_workbook, Workbook
import logging
import numpy as np
import os

#Define Logger

LOGGER = logging.getLogger(__name__)
LOGGER.setLevel(logging.DEBUG)
f = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

""" 
10  DEBUG       logging.debug()     Lowest level. Used to record simple details.
20  INFO        logging.info()      Record general information.
30  WARNING     logging.warning()   Potential issues which may not cause errors in the future.
40  ERROR       logging.error()     Record erros which causes a section of the code to fail.
50  CRITICAL    logging.critical()  Highest level. Blockers which fails your whole program.
 """

def read_excel_vehicular(excel_path
               ) -> type(np.array) and type(np.array) and list:
    """Función que obtiene los conteos por giros de vehículos y motocicletas para todos los sentidos.
    """
    
    #Slices definition
    turn_slice = [slice("G12","G21"),
                  slice("M12","M21"),
                  slice("G24","G33"),
                  slice("M24","M33")]
    
    car_slice = slice("K16","T112")
    moto_slice = slice("U16","AD112")

    #Reading number of turns
    hojas = ['N','S','E','O']
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    LOGGER.info(f"Lectura de excel:\n{excel_path}")
    ws = wb['Inicio']

    try:
        quantities = [
            [row[0].value for row in ws[s]].index(None)
            for s in turn_slice
        ]
    except ValueError:
        LOGGER.error(f"Hay algun dato que no es número en la base de datos")

    #Reading data from sheets
    CAR_LIST = []
    MOTO_LIST = []

    for hoja, num in zip(hojas,quantities):
        ws = wb[hoja]

        car_data = [[elem.value for elem in row][:num] for row in ws[car_slice]]
        try:
            CAR = np.array(car_data, dtype="float")
        except ValueError:
            LOGGER.critical(f"Para autos en la hoja '{hoja}' hay datos que no son números")
        CAR = np.nan_to_num(CAR, nan=0.0)
        CAR_MORNING = CAR[26:38,:]
        CAR_EVENING = CAR[48:60,:]
        CAR_NIGHT = CAR[70:82,:]

        moto_data = [[elem.value for elem in row][:num] for row in ws[moto_slice]]
        try:
            MOTO = np.array(moto_data, dtype="float")
        except ValueError:
            LOGGER.critical(f"Para motos en la hoja '{hoja}' hay datos que no son números")
        MOTO = np.nan_to_num(MOTO, nan=0.0)
        MOTO_MORNING = MOTO[27:39,:]
        MOTO_EVENING = MOTO[49:61,:]
        MOTO_NIGHT = MOTO[71:83,:]

        CAR_LIST.append([CAR_MORNING,CAR_EVENING,CAR_NIGHT]) #POR HOJA
        MOTO_LIST.append([MOTO_MORNING,MOTO_EVENING,MOTO_NIGHT]) #POR HOJA

    wb.close()

    return CAR_LIST, MOTO_LIST

def read_excel_pedestrian(excel_path
               ) -> type(np.array) and type(np.array) and list:
    """Función que obtiene los conteos por giros de vehículos y motocicletas para todos los sentidos.
    """
    
    #Slices definition
    turn_slice = [slice("G13","G22"),
                  slice("K13","K22"),
                  slice("G25","G34"),
                  slice("K25","K34")]
    
    matrix_slice = slice("L20","UY83")
    moto_slice = slice("U16","AD112")

    #Reading number of turns

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    LOGGER.info(f"Lectura de excel:\n{excel_path}")
    ws = wb['Inicio']

    num_giro_i = [0,0,0,0]

    for j, s in enumerate(turn_slice):
        aux = []
        for row in ws[s]:
            aux.append(row[0].value)
        LOGGER.info(f"Número de giros: {aux}")
        try:
            quant = aux.index(None)
        except ValueError: #TODO: Falta analizar el caso en que haya una letra allí!
            quant = len(aux)
        num_giro_i[j] = quant
    
    ws = wb['Data Peatonal']

    #Reading data from sheets
    list_A = [[cell.value for cell in row] for row in ws[matrix_slice]]

    for i, row in enumerate(list_A):
        for j, value in enumerate(row):
            if isinstance(value, str):
                LOGGER.error(f"Row: {i+19}, Column: {j+11}, Value: {value}")

    A = np.array(list_A, dtype="float")
    A[np.isnan(A)] = 0

    morning = []
    evening = []
    night   = []

    INCREMENTOS = [0,140,280,420]

    for incremento, giros in zip(INCREMENTOS, num_giro_i):
        morning.append(
            np.array(
                [
                    A[2:14, (10*ped_type+incremento):(10*ped_type+giros+incremento)]
                    for ped_type in range(4) #Kid, Adult, Old, WC
                ]
            )
        )
        evening.append(
            np.array(
                [
                    A[24:36, (10*ped_type+incremento):(10*ped_type+giros+incremento)]
                    for ped_type in range(4) #Kid, Adult, Old, WC
                ]
            )
        )
        night.append(
            np.array(
                [
                    A[46:58, (10*ped_type+incremento):(10*ped_type+giros+incremento)]
                    for ped_type in range(4) #Kid, Adult, Old, WC
                ]
            )
        )

    wb.close()

    MORNING = []
    EVENING = []
    NIGHT = []

    for morning_i, evening_i, night_i in zip(morning, evening, night):
        morning_concatenated = np.concatenate(morning_i, axis = 1)
        evening_concatenated = np.concatenate(evening_i, axis = 1)
        night_concatenated = np.concatenate(night_i, axis = 1)
        MORNING.append(morning_concatenated)
        EVENING.append(evening_concatenated)
        NIGHT.append(night_concatenated)

    MORNING = np.concatenate(MORNING, axis = 1)
    EVENING = np.concatenate(EVENING, axis = 1)
    NIGHT = np.concatenate(NIGHT, axis = 1)
    return MORNING, EVENING, NIGHT

def find_duplicate_cars(CARS, length, ws, accum_count, sheet, excel_name) -> int:
    """Función para encontrar los patrones repetidos según la longitud del patrón especificado.
    """
    hojas = ['N','S','E','O']
    
    count = accum_count
    repes_list = []
    CAR = np.concatenate((CARS[0],CARS[1]))
    CAR = np.concatenate((CAR,CARS[2]))

    for k in range(CAR.shape[1]):
        GIRO = CAR[:,k]
        for i in range(len(GIRO)-length+1):
            set = GIRO[i:i+length]
            if np.array_equal(set, np.zeros(length)):
                continue
            for j in range(i+1, len(GIRO) - length + 1):
                if np.array_equal(GIRO[j:j+length], set):
                    if sum(set)<=4:
                        continue
                    result = ""
                    for elem in set:
                        result += str(elem)+ ', '
                    ws.cell(row=count+2, column=1, value=result[:-2])
                    ws.cell(row=count+2, column=2, value=hojas[sheet])
                    ws.cell(row=count+2, column=3, value=excel_name)
                    #print(f"Conjunto repetido: {set}, Hoja: {hojas[sheet]}, Fila: {count+2}")
                    repes_list.append(set)
                    count += 1

    return count

def find_duplicate_pedestrian(MORNING, EVENING, NIGHT, length, ws, accum_count, excel_name) -> int:
    count = accum_count

    PEDESTRIAN = np.concatenate((MORNING, EVENING))
    PEDESTRIAN = np.concatenate((PEDESTRIAN, NIGHT))

    for k in range(PEDESTRIAN.shape[1]):
        GIRO = PEDESTRIAN[:,k]
        for i in range(len(GIRO)-length+1):
            set = GIRO[i:i+length]
            if np.array_equal(set,np.zeros(length)): #Delete zero cases
                continue
            for j in range(i+1, len(GIRO) - length + 1):
                if np.array_equal(GIRO[j:j+length], set):
                    if sum(set) <= 4: #No considering low duplicates
                        continue
                    result = ""
                    for elem in set:
                        result += str(elem) + ','
                    ws.cell(row=count+2, column=1, value=result[:-2])
                    ws.cell(row=count+2, column=2, value=excel_name)
                    count += 1

    return count

def jump_single():
    pass

def jump_multiple():
    pass

def main():
    directory = r"C:\Users\dacan\OneDrive\Desktop\PRUEBAS\Supervisor\Entregable Nro 06\7.- Informacion de Campo\Peatonal\Tipico\Pruebitas"
    logger_path = os.path.join(directory,"LOGS")
    if not os.path.exists(logger_path):
        os.mkdir(logger_path)
    fh = logging.FileHandler(os.path.join(directory,"LOGS","means.log"))
    fh.setFormatter(f)
    LOGGER.addHandler(fh)
    excels = os.listdir(directory)
    excels = [excel for excel in excels if excel.endswith('.xlsm') and '~$' not in excel]
    new_excel = os.path.join(directory, "Pattern_Summary.xlsx")
    wb = Workbook()
    wb.save(new_excel)
    wb.close()

    wb = load_workbook(new_excel)
    ws = wb['Sheet']

    ws.cell(row=1, column=1, value="Patrón")
    ws.cell(row=1, column=2, value="Excel")

    accum_count = 0

    for excel in excels:
        MORNING, EVENING, NIGHT = read_excel_pedestrian(os.path.join(directory,excel))
        current_count = find_duplicate_pedestrian(MORNING, EVENING, NIGHT, 4, ws, accum_count, excel)
        accum_count = current_count

    wb.save(new_excel)
    wb.close()
        
    # directory = r"C:\Users\dacan\OneDrive\Desktop\PRUEBAS\Supervisor\Entregable Nro 06\7.- Informacion de Campo\Vehicular\Raw_Data"

    # logger_path = os.path.join(directory,"LOGS")
    # if not os.path.exists(logger_path):
    #     os.mkdir(logger_path)
    # fh = logging.FileHandler(os.path.join(directory,"LOGS","means.log"))
    # fh.setFormatter(f)
    # LOGGER.addHandler(fh)
    # excels = os.listdir(directory)
    # excels = [excel for excel in excels if excel.endswith('.xlsm') and '~$' not in excel]
    # new_excel = os.path.join(directory,"Pattern_Summary.xlsx")
    # if not os.path.exists(new_excel):
    #     wb = Workbook()
    #     wb.save(new_excel)
    #     wb.close()

    # wb = load_workbook(new_excel)
    # ws = wb['Sheet']

    # #Encabezados:
    # ws.cell(row=1, column=1, value="Patrón")
    # ws.cell(row=1, column=2, value="Hoja")
    # ws.cell(row=1, column=3, value="Excel")

    # acumm_count = 0
    # for count, excel in enumerate(excels):
    #     print(f"{count+1}/{len(excels)} Reading: {excel}")
    #     CARS_LIST, MOTOS_LIST = read_excel_vehicular(os.path.join(directory, excel))
    #     for sheet, CARS in enumerate(CARS_LIST): #CARS = Arreglo por sentido.
    #         current_count = find_duplicate(CARS,4,ws,acumm_count, sheet,excel)
    #         acumm_count = current_count
    #     for sheet, MOTOS in enumerate(MOTOS_LIST):
    #         current_count = find_duplicate(MOTOS,4,ws,acumm_count,sheet,excel)
    #         acumm_count = current_count

    # wb.save(new_excel)
    # wb.close()

if __name__ == '__main__':
    main()