from openpyxl import load_workbook
import numpy as np
import os
from collections import Counter

def read_excel_peatonal(excel_path, us): #define la lectura del excel, colocación a 0 de type flotante
    #global hojas #invoca variables fuera de la función
    wb = load_workbook(excel_path,read_only=True,data_only=True)
    data =[]
    ws = wb["Data Peatonal"]
    filas = ws["L19:UY19"]
    valores = [celda.value for fil in filas for celda in fil]
    data_hoja = [[celda.value for celda in fila] for fila in ws['L20:UY79']] #fila-celda SE CAMBIA EL RANGO
    try:
        matriz_numpy = np.array(data_hoja,dtype="float") #conversión a np
        if us:
            index = (np.where(~np.isnan(matriz_numpy).all(axis=1))[0])
            if len(index) > 0:
                quarter_hour = index[0]
            else:
                quarter_hour = 0 #No existe en teoría
        else:
            quarter_hour = 0
    except TypeError:
        print("Hay un valor que no es número en la base de datos")
        _, excel_name = os.path.split(excel_path)
        print(f"Excel = {excel_name}")
        print(f"Hoja = Data Peatonal")
    matriz_numpy[np.isnan(matriz_numpy)] = 0 #0 0 a vacios
    data.append(matriz_numpy)

    ws = wb["Inicio"] #guarda fecha
    fecha = str(ws['G5'].value) #Por cada excel

    return data,fecha,valores, quarter_hour

def read_excel_vehicular(excel_path,us):
    hojas = ["N","S","E","O"]
    wb = load_workbook(excel_path,read_only=True,data_only=True)
    data=[]
    val =[]
    quarter_hour = []
    for hoja in hojas:
        ws = wb[hoja]
        filas = ws["K15:HB15"]
        valores = [celda.value for fil in filas for celda in fil] #Encabezado
        val.append(valores)
        data_hoja = [[celda.value for celda in fila] for fila in ws['K16:HB111']] #fila-celda
        try:
            matriz_numpy = np.array(data_hoja,dtype="float") #conversión a np
            if us:
                index = (np.where(~np.isnan(matriz_numpy).all(axis=1))[0])
                if len(index)>0:
                    quarter_hour.append(index[0])
                else:
                    quarter_hour.append(0)
            else:
                quarter_hour.append(0)
        except ValueError:
            print("Hay un valor que no es número en la base de datos")
            _,excel_name = os.path.split(excel_path)
            print(f"Excel = {excel_name}")
            print(f"Hoja = {hoja}")
            return print("^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^\n################### PROCESO DETENIDO - ERROR #####################")
        matriz_numpy[np.isnan(matriz_numpy)] = 0 #0 0 a vacios
        data.append(matriz_numpy)

    ws = wb["Inicio"] #guarda fecha
    fecha = str(ws['G8'].value) #Por cada excel
    if us: #True = Protransito
        positive_numbers = [num for num in quarter_hour if num > 0]
        if positive_numbers:
            contador = Counter(positive_numbers)
            common = contador.most_common(1)[0][0]
        else:
            common = 0
        return data,fecha,val,common
    else:
        common = 0
    return data,fecha,val,common